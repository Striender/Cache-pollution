#!/usr/bin/env python3
import os
import re
import pandas as pd
from collections import defaultdict
import json
import shutil
import copy

# Try to import openpyxl and guide the user if it's not installed.
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import MergedCell
except ImportError:
    print("The 'openpyxl' library is required to write formatted Excel files.")
    print("Please install it on your server by running: pip install openpyxl")
    exit()

# ANSI escape codes for terminal colors
class TColors:
    OKGREEN = '\033[92m'
    ENDC = '\033[0m'

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]

def load_json_data(file_path):
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, 'r') as f:
            return json.load(f)
    except:
        return {}

def save_json_data(file_path, data):
    try:
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)
    except:
        pass


def parse_champsim_file(filepath):
    metrics = {
        "Trace File": os.path.basename(filepath) if filepath else None,
        "IPC": None,

        # L1D
        "L1D Total Access": None, "L1D Total Hit": None, "L1D Total Miss": None, "L1D Total MPKI": None,
        "L1D Load Miss": None, "L1D Load MPKI": None,
        "L1D Prefetch Access": None,
        "L1D Prefetch Requested": None,
        "L1D Prefetch Issued": None,
        "L1D Prefetch Useful": None,
        "L1D Prefetch Useless": None,
        "L1D Useful Load Prefetches": None,
        "L1D Prefetch Issued To Lower Level": None,
        "L1D Prefetch Accuracy": None,
        "L1D Average Miss Latency": None,
        "L1D Late Prefetches": None,
        "L1D Prefetch Coverage": None,

        # L2C
        "L2C Total Access": None, "L2C Total Hit": None, "L2C Total Miss": None, "L2C Total MPKI": None,
        "L2C Load Miss": None, "L2C Load MPKI": None,
        "L2C Prefetch Access": None,
        "L2C Prefetch Requested": None,
        "L2C Prefetch Issued": None,
        "L2C Prefetch Useful": None,
        "L2C Prefetch Useless": None,
        "L2C Useful Load Prefetches": None,
        "L2C Prefetch Issued To Lower Level": None,
        "L2C Prefetch Accuracy": None,
        "L2C Average Miss Latency": None,
        "L2C Late Prefetches": None,
        "L2C Prefetch Coverage": None,
        "L2C Pollution": None,

        # LLC
        "LLC Total Access": None, "LLC Total Hit": None, "LLC Total Miss": None, "LLC Total MPKI": None,
        "LLC Load Miss": None, "LLC Load MPKI": None,
        "LLC Prefetch Access": None,
        "LLC Prefetch Requested": None,
        "LLC Prefetch Issued": None,
        "LLC Prefetch Useful": None,
        "LLC Prefetch Useless": None,
        "LLC Useful Load Prefetches": None,
        "LLC Prefetch Accuracy": None,
        "LLC Average Miss Latency": None,
        "LLC Late Prefetches": None,
        "LLC Prefetch Coverage": None,
        "LLC Pollution": None,
    }

    if not filepath:
        return list(metrics.keys())

    try:
        with open(filepath, 'r', errors='ignore') as f:
            content = f.read()

            # IPC
            m = re.search(r"CPU 0 cumulative IPC:\s+([\d.]+)", content)
            if m:
                metrics["IPC"] = float(m.group(1))

            # L1D total
            m = re.search(r"L1D TOTAL\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if m:
                metrics["L1D Total Access"] = int(m.group(1))
                metrics["L1D Total Hit"] = int(m.group(2))
                metrics["L1D Total Miss"] = int(m.group(3))
                metrics["L1D Total MPKI"] = float(m.group(4))

            # L1D load
            m = re.search(r"L1D LOAD\s+ACCESS:\s+\d+\s+HIT:\s+\d+\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if m:
                metrics["L1D Load Miss"] = int(m.group(1))
                metrics["L1D Load MPKI"] = float(m.group(2))

            # L1D prefetch access
            m = re.search(r"L1D PREFETCH\s+ACCESS:\s+(\d+)", content)
            if m:
                metrics["L1D Prefetch Access"] = int(m.group(1))

            # UPDATED L1D REQUESTED / ISSUED / USEFUL / USELESS
            m = re.search(r"L1D PREFETCH\s+REQUESTED:\s+(\d+)\s+ISSUED:\s+(\d+)\s+USEFUL:\s+(\d+)\s+USELESS:\s+(\d+)", content)
            if m:
                metrics["L1D Prefetch Requested"] = int(m.group(1))
                metrics["L1D Prefetch Issued"] = int(m.group(2))
                metrics["L1D Prefetch Useful"] = int(m.group(3))
                metrics["L1D Prefetch Useless"] = int(m.group(4))

            m = re.search(r"L1D USEFUL LOAD PREFETCHES:\s+(\d+)", content)
            if m:
                metrics["L1D Useful Load Prefetches"] = int(m.group(1))

            # NEW: L1D PREFETCH ISSUED TO LOWER LEVEL
            m = re.search(r"L1D USEFUL LOAD PREFETCHES:\s+\d+\s+PREFETCH ISSUED TO LOWER LEVEL:\s+(\d+)", content)
            if m:
                metrics["L1D Prefetch Issued To Lower Level"] = int(m.group(1))

            m = re.search(r"L1D USEFUL LOAD PREFETCHES:.*?ACCURACY:\s+([\d.inf-]+)", content)
            if m:
                try:
                    metrics["L1D Prefetch Accuracy"] = float(m.group(1))
                except:
                    metrics["L1D Prefetch Accuracy"] = m.group(1)

            m = re.search(r"L1D AVERAGE MISS LATENCY:\s+([\d.]+)", content)
            if m:
                metrics["L1D Average Miss Latency"] = float(m.group(1))

            m = re.search(r"L1D TIMELY PREFETCHES:\s+\d+\s+LATE PREFETCHES:\s+(\d+)", content)
            if m:
                metrics["L1D Late Prefetches"] = int(m.group(1))


            # ----------------- L2C -------------------
            m = re.search(r"L2C TOTAL\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if m:
                metrics["L2C Total Access"] = int(m.group(1))
                metrics["L2C Total Hit"] = int(m.group(2))
                metrics["L2C Total Miss"] = int(m.group(3))
                metrics["L2C Total MPKI"] = float(m.group(4))

            m = re.search(r"L2C LOAD\s+ACCESS:\s+\d+\s+HIT:\s+\d+\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if m:
                metrics["L2C Load Miss"] = int(m.group(1))
                metrics["L2C Load MPKI"] = float(m.group(2))

            m = re.search(r"L2C PREFETCH\s+ACCESS:\s+(\d+)", content)
            if m:
                metrics["L2C Prefetch Access"] = int(m.group(1))

            # UPDATED L2C REQUESTED / ISSUED / USEFUL / USELESS
            m = re.search(r"L2C PREFETCH\s+REQUESTED:\s+(\d+)\s+ISSUED:\s+(\d+)\s+USEFUL:\s+(\d+)\s+USELESS:\s+(\d+)", content)
            if m:
                metrics["L2C Prefetch Requested"] = int(m.group(1))
                metrics["L2C Prefetch Issued"] = int(m.group(2))
                metrics["L2C Prefetch Useful"] = int(m.group(3))
                metrics["L2C Prefetch Useless"] = int(m.group(4))

            m = re.search(r"L2C USEFUL LOAD PREFETCHES:\s+(\d+)", content)
            if m:
                metrics["L2C Useful Load Prefetches"] = int(m.group(1))

            # NEW: L2C PREFETCH ISSUED TO LOWER LEVEL
            m = re.search(r"L2C USEFUL LOAD PREFETCHES:\s+\d+\s+PREFETCH ISSUED TO LOWER LEVEL:\s+(\d+)", content)
            if m:
                metrics["L2C Prefetch Issued To Lower Level"] = int(m.group(1))

            m = re.search(r"L2C USEFUL LOAD PREFETCHES:.*?ACCURACY:\s+([\d.inf-]+)", content)
            if m:
                try:
                    metrics["L2C Prefetch Accuracy"] = float(m.group(1))
                except:
                    metrics["L2C Prefetch Accuracy"] = m.group(1)

            m = re.search(r"L2C AVERAGE MISS LATENCY:\s+([\d.]+)", content)
            if m:
                metrics["L2C Average Miss Latency"] = float(m.group(1))

            m = re.search(r"L2C TIMELY PREFETCHES:\s+\d+\s+LATE PREFETCHES:\s+(\d+)", content)
            if m:
                metrics["L2C Late Prefetches"] = int(m.group(1))

            m = re.search(r"Total pollution count in L2\s*:\s+([\d.]+)", content)
            if m:
                metrics["L2C Pollution"] = float(m.group(1))


            # ----------------- LLC -------------------
            m = re.search(r"LLC TOTAL\s+ACCESS:\s+(\d+)\s+HIT:\s+(\d+)\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if m:
                metrics["LLC Total Access"] = int(m.group(1))
                metrics["LLC Total Hit"] = int(m.group(2))
                metrics["LLC Total Miss"] = int(m.group(3))
                metrics["LLC Total MPKI"] = float(m.group(4))

            m = re.search(r"LLC LOAD\s+ACCESS:\s+\d+\s+HIT:\s+\d+\s+MISS:\s+(\d+).*?MPKI:\s+([\d.]+)", content)
            if m:
                metrics["LLC Load Miss"] = int(m.group(1))
                metrics["LLC Load MPKI"] = float(m.group(2))

            m = re.search(r"LLC PREFETCH\s+ACCESS:\s+(\d+)", content)
            if m:
                metrics["LLC Prefetch Access"] = int(m.group(1))

            # UPDATED LLC REQUESTED / ISSUED / USEFUL / USELESS
            m = re.search(r"LLC PREFETCH\s+REQUESTED:\s+(\d+)\s+ISSUED:\s+(\d+)\s+USEFUL:\s+(\d+)\s+USELESS:\s+(\d+)", content)
            if m:
                metrics["LLC Prefetch Requested"] = int(m.group(1))
                metrics["LLC Prefetch Issued"] = int(m.group(2))
                metrics["LLC Prefetch Useful"] = int(m.group(3))
                metrics["LLC Prefetch Useless"] = int(m.group(4))

            m = re.search(r"LLC USEFUL LOAD PREFETCHES:\s+(\d+)", content)
            if m:
                metrics["LLC Useful Load Prefetches"] = int(m.group(1))

            m = re.search(r"LLC USEFUL LOAD PREFETCHES:.*?ACCURACY:\s+([\d.inf-]+)", content)
            if m:
                try:
                    metrics["LLC Prefetch Accuracy"] = float(m.group(1))
                except:
                    metrics["LLC Prefetch Accuracy"] = m.group(1)

            m = re.search(r"LLC AVERAGE MISS LATENCY:\s+([\d.]+)", content)
            if m:
                metrics["LLC Average Miss Latency"] = float(m.group(1))

            m = re.search(r"LLC TIMELY PREFETCHES:\s+\d+\s+LATE PREFETCHES:\s+(\d+)", content)
            if m:
                metrics["LLC Late Prefetches"] = int(m.group(1))

            m = re.search(r"Total pollution count in LLC:\s+([\d.]+)", content)
            if m:
                metrics["LLC Pollution"] = float(m.group(1))

    except IOError:
        return None

    return metrics


def apply_border_to_range(worksheet, row_range, col_range, border_style):
    """Helper function to apply a border to a range of cells."""
    for row in worksheet.iter_rows(min_row=row_range[0], max_row=row_range[1],
                                   min_col=col_range[0], max_col=col_range[1]):
        for cell in row:
            cell.border = border_style

def main():
    """
    Main function to find ChampSim files, parse them, and save them to a single
    formatted Excel file with multiple sheets, preserving user-added sheets.
    """
    # --- CONFIGURATION ---
    RESULTS_DIR = "../results/results/"
    OUTPUT_DIR = "/home2/neeraj/OneDrive/Research_Data"
    EXCEL_OUTPUT_FILE = "data_dump_issued.xlsx"
    PROCESSED_LOG_FILE = os.path.join(OUTPUT_DIR, ".processed_files.log")
    DATA_CACHE_FILE = os.path.join(OUTPUT_DIR, ".data_cache.json")
    # -------------------

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.isdir(RESULTS_DIR):
        print(f"Error: Directory '{RESULTS_DIR}' not found.")
        return

    processed_files_log = load_json_data(PROCESSED_LOG_FILE)
    cached_data = load_json_data(DATA_CACHE_FILE)
    
    # The structure is {group_key: {filepath: metrics}}
    # We will rebuild this dict completely to handle file deletions properly
    data_by_prefetcher = defaultdict(dict)
    
    new_files_count = 0
    skipped_files_count = 0
    announced_dirs = set()  # To prevent printing the same directory multiple times
    
    print(f"Starting scan in directory: '{RESULTS_DIR}'...")
    # Walk through the directory tree to collect all data
    for root, dirs, files in os.walk(RESULTS_DIR):
        if not files:
            continue

        relative_path = os.path.relpath(root, RESULTS_DIR)
        path_parts = relative_path.split(os.sep)

        group_key, experiment = None, None

        if len(path_parts) == 3:  # Standard case: results/pref_l1/berti/exp1
            cache_level, prefetcher, experiment = path_parts
            group_key = f"{cache_level}_{prefetcher}"
        elif len(path_parts) == 2:  # Edge case for no_pref: results/no_pref/exp1
            cache_level, experiment = path_parts
            if cache_level == 'no_pref':
                group_key = cache_level
        
        if group_key and experiment:
            for filename in sorted(files):
                filepath = os.path.join(root, filename)
                
                file_mod_time = os.path.getmtime(filepath)
                # If file is unchanged, load its data from cache instead of re-parsing
                if processed_files_log.get(filepath) == file_mod_time:
                    if group_key in cached_data and filepath in cached_data[group_key]:
                        data_by_prefetcher[group_key][filepath] = cached_data[group_key][filepath]
                    skipped_files_count += 1
                    continue
                
                # If file is new or modified, announce the directory once
                if root not in announced_dirs:
                    print(f"{TColors.OKGREEN}Processing new/modified files in: {relative_path}{TColors.ENDC}")
                    announced_dirs.add(root)

                new_files_count += 1
                metrics = parse_champsim_file(filepath)
                if metrics:
                    metrics["Experiment"] = experiment
                    data_by_prefetcher[group_key][filepath] = metrics
                    processed_files_log[filepath] = file_mod_time

    print(f"\nScan complete. Found {new_files_count} new/modified files. Skipped {skipped_files_count} unchanged files.")

    if new_files_count == 0:
        print("\nOutput is already up-to-date.")
        return

    print(f"\nProcessing data and updating {EXCEL_OUTPUT_FILE}...")
    final_output_path = os.path.join(OUTPUT_DIR, EXCEL_OUTPUT_FILE)
    temp_output_path = os.path.join("/tmp", f"{os.getpid()}_{EXCEL_OUTPUT_FILE}")  # Unique temp file name
    
    try:
        from openpyxl import Workbook
        book = Workbook()
        if 'Sheet' in book.sheetnames:
            # Ensure there's always at least one sheet before removing the default one
            if len(book.sheetnames) > 1:
                book.remove(book.active)
            else:  # If it's the only sheet, just clear it instead of removing
                book.active.title = "Placeholder"  # Rename temporarily if needed
                for row in book.active.iter_rows():
                    for cell in row:
                        cell.value = None

        existing_custom_sheets = {}
        if os.path.exists(final_output_path):
            try:
                old_book = load_workbook(final_output_path)
                for sheet_name in old_book.sheetnames:
                    if not sheet_name.startswith('raw_'):
                        print(f"Preserving your custom sheet: {sheet_name}")
                        existing_custom_sheets[sheet_name] = old_book[sheet_name]

            except Exception as e:
                print(f"Warning: Could not load or copy sheets from existing workbook. It might be corrupted. A new file will be created. Error: {e}")

        # Define styles once
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        main_header_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")  # Peach
        data_header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # Dark Gray
        sub_header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Get the defined headers safely
        headers = parse_champsim_file(None)

        # Build the new workbook content first
        for group_key in sorted(data_by_prefetcher.keys()):
            data_list = list(data_by_prefetcher[group_key].values())
            if not data_list:
                continue

            print(f"Processing group: {group_key}")
            
            df = pd.DataFrame(data_list)
            sheet_name = f"{group_key}"
            
            # Create sheet in the new book
            worksheet = book.create_sheet(title=sheet_name)
            
            # --- Create and Write Main Header ---
            if group_key == 'no_pref':
                main_header_text = "Baseline (No Prefetcher)"
            else:
                parts = group_key.split('_')
                cache_level_str = parts[1].upper() if len(parts) > 1 else ''
                prefetcher_name = '_'.join(parts[2:]).capitalize() if len(parts) > 2 else parts[0].capitalize()
                main_header_text = f"Data Prefetcher: {prefetcher_name} at {cache_level_str}"
            
            num_cols = len(headers)

            worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=num_cols)
            main_header_cell = worksheet.cell(row=1, column=1, value=main_header_text)
            main_header_cell.font = Font(bold=True, size=14)
            main_header_cell.alignment = center_alignment
            main_header_cell.fill = main_header_fill
            apply_border_to_range(worksheet, (1, 2), (1, num_cols), thin_border)

            # Write and Style the data column headers on row 3
            for col_num, col_name in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col_num, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = data_header_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            worksheet.row_dimensions[3].height = 30
            
            worksheet.freeze_panes = 'A4'  # Freeze rows 1, 2, and 3

            experiments = sorted(df['Experiment'].unique(), key=natural_sort_key)
            current_row = 3

            for experiment in experiments:
                df_experiment = df[df['Experiment'] == experiment].copy()
                
                if current_row > 3:
                    current_row += 1 

                bold_font = Font(bold=True, size=12)
                
                # --- Create Descriptive Experiment Header ---
                try:
                    exp_parts = experiment.split('_')
                    exp_num = ''.join(filter(str.isdigit, exp_parts[0]))
                    l2_policy = exp_parts[1].upper()
                    llc_policy = exp_parts[2].upper()
                    exp_header_text = f"Experiment {exp_num}: Replacement Policy {l2_policy} at L2 and {llc_policy} at LLC"
                except (IndexError, ValueError):
                    exp_header_text = experiment.replace('_', ' ').title()

                worksheet.merge_cells(start_row=current_row + 1, start_column=1,
                                      end_row=current_row + 1, end_column=num_cols)
                
                header_cell = worksheet.cell(row=current_row + 1, column=1, value=exp_header_text)
                header_cell.font = bold_font
                header_cell.alignment = center_alignment
                header_cell.fill = sub_header_fill
                apply_border_to_range(worksheet, (current_row + 1, current_row + 1), (1, num_cols), thin_border)
                worksheet.row_dimensions[current_row + 1].height = 30

                # Reindex to ensure all columns are present, then fill missing with 'NaN'
                df_to_write = df_experiment.reindex(columns=headers).fillna('NaN')
                
                # Write data using openpyxl to apply styles cell by cell
                for r_idx, row_data in enumerate(df_to_write.itertuples(index=False), start=current_row + 2):
                    for c_idx, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                        # First column is left-aligned, others are right-aligned
                        if c_idx == 1:
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = right_alignment
                
                # --- Add Coverage Formulas ---
                # Get column letters for the calculation
                l1d_useful_col = get_column_letter(headers.index("L1D Useful Load Prefetches") + 1)
                l1d_load_miss_col = get_column_letter(headers.index("L1D Load Miss") + 1)
                l1d_coverage_col = get_column_letter(headers.index("L1D Prefetch Coverage") + 1)
                
                l2c_useful_col = get_column_letter(headers.index("L2C Useful Load Prefetches") + 1)
                l2c_load_miss_col = get_column_letter(headers.index("L2C Load Miss") + 1)
                l2c_coverage_col = get_column_letter(headers.index("L2C Prefetch Coverage") + 1)

                llc_useful_col = get_column_letter(headers.index("LLC Useful Load Prefetches") + 1)
                llc_load_miss_col = get_column_letter(headers.index("LLC Load Miss") + 1)
                llc_coverage_col = get_column_letter(headers.index("LLC Prefetch Coverage") + 1)

                for r_idx in range(current_row + 2, current_row + 2 + len(df_to_write)):
                    # L1D Coverage Formula
                    l1d_denominator = f"({l1d_useful_col}{r_idx}+{l1d_load_miss_col}{r_idx})"
                    worksheet[f"{l1d_coverage_col}{r_idx}"] = (
                        f'=IF({l1d_denominator}=0, 0, {l1d_useful_col}{r_idx}/{l1d_denominator})'
                    )

                    # L2C Coverage Formula
                    l2c_denominator = f"({l2c_useful_col}{r_idx}+{l2c_load_miss_col}{r_idx})"
                    worksheet[f"{l2c_coverage_col}{r_idx}"] = (
                        f'=IF({l2c_denominator}=0, 0, {l2c_useful_col}{r_idx}/{l2c_denominator})'
                    )

                    # LLC Coverage Formula
                    llc_denominator = f"({llc_useful_col}{r_idx}+{llc_load_miss_col}{r_idx})"
                    worksheet[f"{llc_coverage_col}{r_idx}"] = (
                        f'=IF({llc_denominator}=0, 0, {llc_useful_col}{r_idx}/{llc_denominator})'
                    )

                current_row += 1 + len(df_experiment)
            
            bold_font_for_trace = Font(bold=True)
            for cell in worksheet['A']:
                if cell.row > 3 and cell.value and worksheet.cell(row=cell.row, column=2).value:
                    cell.font = bold_font_for_trace

            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in worksheet[column_letter]:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value:
                        max_length = max(len(str(cell.value)), max_length)
                worksheet.column_dimensions[column_letter].width = max_length + 2
            
            print(f" -> Finished processing sheet: {sheet_name}")

        # Add the preserved custom sheets back into the workbook
        for sheet_name, old_ws in existing_custom_sheets.items():
            new_ws = book.create_sheet(title=sheet_name)
            for row in old_ws.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value
                    if cell.has_style:
                        new_ws[cell.coordinate].font = cell.font.copy()
                        new_ws[cell.coordinate].border = cell.border.copy()
                        new_ws[cell.coordinate].fill = cell.fill.copy()
                        new_ws[cell.coordinate].number_format = cell.number_format
                        new_ws[cell.coordinate].protection = cell.protection.copy()
                        new_ws[cell.coordinate].alignment = cell.alignment.copy()

        # Save the entire workbook to the temporary path first
        book.save(temp_output_path)

        # Move the completed file to the final destination
        shutil.move(temp_output_path, final_output_path)
        
        # Clean the cache for saving by removing internal flags
        for group in data_by_prefetcher.values():
            for record in group.values():
                record.pop('_is_new', None)
                
        save_json_data(DATA_CACHE_FILE, data_by_prefetcher)
        save_json_data(PROCESSED_LOG_FILE, processed_files_log)
        print(f"\nSuccessfully created/updated Excel file: {final_output_path}")

    except Exception as e:
        print(f"\nAn error occurred while writing the Excel file: {e}")
        # Clean up the temporary file if it exists and writing failed
        if os.path.exists(temp_output_path):
            try:
                os.remove(temp_output_path)
            except OSError:
                pass


if __name__ == "__main__":
    main()
